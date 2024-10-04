
from openpyxl import Workbook
from datetime import datetime

wb = Workbook()

ws = wb.active

ws.title = "livevox_sheet1"

ws["B9"] = "Hello World"
ws["D9"] = 85297000
ws["F9"] = datetime.now()

wb.save("MyFirstExcel.xlsx")
