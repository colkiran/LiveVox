
from openpyxl import load_workbook
from prettytable import PrettyTable

wb = load_workbook("MyFirstExcel.xlsx")

wb.active = wb['Players']

ws = wb.active

print(ws.dimensions)

datarange = ws[ws.dimensions]

row = 1
for c1, c2, c3, c4, c5 in datarange:
    if row == 1:
        plyrTbl = PrettyTable([c1.value, c2.value, c3.value, c4.value, c5.value])
    else:
        plyrTbl.add_row([c1.value, c2.value, c3.value, c4.value, c5.value])
    row += 1

print(plyrTbl)