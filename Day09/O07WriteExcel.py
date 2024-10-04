
from openpyxl import load_workbook

wb = load_workbook("MyFirstExcel.xlsx")

wb.create_sheet("Players")

wb.active = wb["Players"]

ws = wb.active

cell = ws["A5"]

data = [
    ['PlayerName', 'Age', "Runs", "Matches", "Country"],
    ['Sachin', 31, 5812, 950, 'India'],
    ['Lara', 33, 34850, 685, 'West Indies'],
    ['Ponting', 35, 78299,580, "Australia"],
    ['Jayasurya', 36, 2625, 389, "Srilanka"],
    ['Inzamam', 32, 23407, 450, "Pakistan"]
]

for row in data:
    ws.append(row)

wb.save("MyFirstExcel.xlsx")
